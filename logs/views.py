from django.shortcuts import render
# 引入Student的类
from .models import Cms_log
from .models import Cms_users
# 引入JsonResponse模块
from django.http import JsonResponse
# 导入json模块
import json
# 导入Q查询
from django.db.models import Q
import time
import random

from django.db import connection


# 导入uuid类
import uuid
# 导入哈希库
import hashlib
# 导入Setting
from django.conf import settings
# 导入os
import os
# 引入处理Excel模块
import openpyxl
# Create your views here.
from django.db import connection
import numpy as np

# Create your views here.
# 生成随机token

# token生成模块
from django.db.models import Count



#根据前端的传参条件拼接Q
def query_q(dict):
    enddict1 ={}
    enddict2 ={}
    if dict['inputStr'] :
        enddict1['ownername__icontains'] = dict['inputStr']
        enddict1['deptname__icontains'] = dict['inputStr']
        enddict1['event_system__icontains'] = dict['inputStr']
        enddict1['event_mark__icontains'] = dict['inputStr']
        enddict1['event_type__icontains'] = dict['inputStr']
        enddict1['resolvent__icontains'] = dict['inputStr']
        enddict1['event_from__icontains'] = dict['inputStr']
        enddict1['handler__icontains'] = dict['inputStr']
        enddict1['proposer__icontains'] = dict['inputStr']
        enddict1['status__icontains'] = dict['inputStr']
        enddict1['mark__icontains'] = dict['inputStr']

    if dict['handler'] :
        enddict2['handler'] = dict['handler']

    if dict['admin'] =='10':
        del enddict2['handler']
    else:
        enddict2['handler'] = dict['handler']

    if dict['inputBegdate']:
        enddict2['createdate__gte'] = dict['inputBegdate']


    if dict['inputEnddate']:
        enddict2['createdate__lte'] = dict['inputEnddate']

    con = Q() #汇总的Q

    q1 = Q() #普通字段汇总
    q1.connector = 'OR'
    for i in enddict1:
        q1.add(Q(**{i: enddict1[i]}), Q.OR)

    q2 = Q()#特殊字段汇总
    q2.connector = 'AND'
    for i in enddict2:
        q2.add(Q(**{i: enddict2[i]}), Q.AND)

   #最终汇总
    con.add(q1, 'AND')
    con.add(q2, 'AND')

    return  con


def generate_token(name):
    c_time = str(time.time())
    r = str(random.random())
    return hashlib.new("md5", (c_time + r + name).encode("utf-8")).hexdigest()

# sha1密码加密


def sha1_password(res: str):
    import hashlib
    """
    使用sha1加密算法，返回str加密后的字符串
    """
    sha = hashlib.sha1(res.encode('utf-8'))
    encrypts = sha.hexdigest()
    return encrypts


# 获取所有日志信息


def get_cmslogs(request):
    data = json.loads(request.body.decode('utf-8'))

    """获取所有日志信息"""
    try:
        # 为管理员查所有
        if data['admin'] == '10':
            obj_cmslog = Cms_log.objects.all().values()

        else:
            # 只查自己的
            obj_cmslog = Cms_log.objects.filter(
                Q(handler=data['handler'])).values()
            # 把外层的容器转为List
            # 返回add by 20210330


        cmslog = list(obj_cmslog)

        for iii in obj_cmslog:
            iii['createdate'] = iii['createdate'].strftime('%Y-%m-%d')
            if iii['resolventdate']:
                iii['resolventdate'] = iii['resolventdate'].strftime('%Y-%m-%d')
        global export_list
        export_list=list(obj_cmslog)


        return JsonResponse({'code': 1, 'data': cmslog})
    except Exception as e:
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "获取日志时出现异常，具体错误：" + str(e)})

# 根据条件日志信息


def query_cmslogs(request):
    """查询日志信息"""
    # 接收传递过来的查询条件--- axios默认是json --- 字典类型（'inputStr'）-- data['inputStr']

    data = json.loads(request.body.decode('utf-8'))

    # print(obj_cmslog.query.__str__())
    sql=query_q(data)

    try:
        obj_cmslog = Cms_log.objects.filter(sql).values()
        # 返回add by 20210330
        cmslog = list(obj_cmslog)
        for iii in obj_cmslog:
            iii['createdate'] = iii['createdate'].strftime('%Y-%m-%d')
            if iii['resolventdate']:
                iii['resolventdate'] = iii['resolventdate'].strftime('%Y-%m-%d')
        global export_list
        export_list = list(obj_cmslog)
        return JsonResponse({'code': 1, 'data': cmslog})
    except Exception as e:
        # 如果出现异常，返回
         return JsonResponse(
            {'code': 0, 'msg': "查询日志信息时出现异常，具体错误：" + str(e)})






    #
    # if data['inputBegdate'] and data['inputEnddate']:
    #     try:
    #         if data['admin'] == '10':
    #
    #             obj_cmslog = Cms_log.objects.filter(Q(createdate__gte=data['inputBegdate'])  # gte：大于等于某个时间
    #                                                 # lte：小于等于
    #                                                 & Q(createdate__lte=data['inputEnddate']),
    #                                                 Q(id__icontains=data['inputStr'])
    #                                                 | Q(ownername__icontains=data['inputStr'])
    #                                                 | Q(deptname__icontains=data['inputStr'])
    #                                                 | Q(event_system__icontains=data['inputStr'])
    #                                                 | Q(event_mark__icontains=data['inputStr'])
    #                                                 | Q(event_type__icontains=data['inputStr'])
    #                                                 | Q(resolvent__icontains=data['inputStr'])
    #                                                 | Q(event_from__icontains=data['inputStr'])
    #                                                 | Q(handler__icontains=data['inputStr'])
    #                                                 | Q(proposer__icontains=data['inputStr'])
    #                                                 | Q(status__icontains=data['inputStr'])
    #                                                 | Q(mark__icontains=data['inputStr'])).values()
    #
    #
    #         else:
    #
    #             obj_cmslog = Cms_log.objects.filter(Q(createdate__gte=data['inputBegdate'])  # gte：大于等于某个时间
    #                                                 # lte：小于等于
    #                                                 & Q(createdate__lte=data['inputEnddate']) & Q(handler=data['handler']),
    #                                                 Q(id__icontains=data['inputStr'])
    #                                                 | Q(ownername__icontains=data['inputStr'])
    #                                                 | Q(deptname__icontains=data['inputStr'])
    #                                                 | Q(event_system__icontains=data['inputStr'])
    #                                                 | Q(event_mark__icontains=data['inputStr'])
    #                                                 | Q(event_type__icontains=data['inputStr'])
    #                                                 | Q(resolvent__icontains=data['inputStr'])
    #                                                 | Q(event_from__icontains=data['inputStr'])
    #                                                 | Q(proposer__icontains=data['inputStr'])
    #                                                 | Q(status__icontains=data['inputStr'])
    #                                                 | Q(mark__icontains=data['inputStr'])).values()
    #
    #
    #         cmslog = list(obj_cmslog)
    #         return JsonResponse({'code': 1, 'data': cmslog})
    #     except Exception as e:
    #         # 如果出现异常，返回
    #         return JsonResponse(
    #             {'code': 0, 'msg': "查询日志信息时出现异常，具体错误：" + str(e)})
    # else:
    #     return JsonResponse({'code': 0, 'msg': "请输入时间后进行查询！"})
    #

# 添加日志到数据库  会返回所有的数据 ？？？


def add_cmslogs(request):
    """添加日志到数据库"""
    # 接收传递过来的查询条件--- axios默认是json --- 字典类型（'inputStr'）-- data['inputStr']
    data = json.loads(request.body.decode('utf-8'))
    print(data)
    if data['resolventdate']=='' or data['resolventdate']=='Invalid date':
        data['resolventdate'] =None
    try:
        # 新建日志对象
        obj_cmslog = Cms_log(id=1,
                             ownername=data['ownername'],
                             deptname=data['deptname'],
                             event_system=data['event_system'],
                             event_mark=data['event_mark'],
                             event_type=data['event_type'],
                             resolvent=data['resolvent'],
                             event_from=data['event_from'],
                             handler=data['handler'],
                             proposer=data['proposer'],
                             createdate=data['createdate'],
                             resolventdate=data['resolventdate'],
                             status=data['status'],
                             mark=data['mark'],)

        obj_cmslog.save()

        # 保存后返回
        # obj_cmslog = Cms_log.objects.all().values()
        # # 把外层的容器转为List
        # cmslog = list(obj_cmslog)
        # # 返回
        return JsonResponse({'code': 1, 'data': '添加日志成功'})

    except Exception as e:
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "添加日志时出现异常，具体错误：" + str(e)})

# 获取待修改的日志内容
def update_getone_cmslogs(request):
    data = json.loads(request.body.decode('utf-8'))
    print(data)
    print(data['id'])
    """获取待修改日志信息"""
    try:
        #
        obj_cmslog = Cms_log.objects.filter(
        id=data['id']).values()

        # 返回


        cmslog = list(obj_cmslog)
        print(cmslog)

        return JsonResponse({'code': 1, 'list': cmslog[0]})
    except Exception as e:
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "获取待修改的日志内容出现异常，具体错误：" + str(e)})

# 根据条件日志信息


def update_cmslogs(request):
    data = json.loads(request.body.decode('utf-8'))
    print(data)
    if data['resolventdate'] == '' or data['resolventdate'] == 'Invalid date':
        data['resolventdate'] = None
    try:
        # 根据id找到需要修改的日志记录
        obj_cmslog = Cms_log.objects.get(id=data['id'])
        # 依次进行修改

        obj_cmslog.ownername = data['ownername']
        obj_cmslog.deptname = data['deptname']
        obj_cmslog.event_system = data['event_system']
        obj_cmslog.event_mark = data['event_mark']
        obj_cmslog.event_type = data['event_type']
        obj_cmslog.resolvent = data['resolvent']
        obj_cmslog.event_from = data['event_from']
        obj_cmslog.handler = data['handler']
        obj_cmslog.proposer = data['proposer']
        obj_cmslog.createdate = data['createdate']
        obj_cmslog.resolventdate = data['resolventdate']
        obj_cmslog.status = data['status']
        obj_cmslog.mark = data['mark']

        obj_cmslog.save()
        print('修改成功')
        # 保存后返回
        # obj_cmslog = Cms_log.objects.all().values()
        # # 把外层的容器转为List
        # cmslog = list(obj_cmslog)
        # 返回
        return JsonResponse({'code': 1, 'msg': '日志修改成功'})

    except Exception as e:
        print(e)
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "修改日志时出现异常，具体错误：" + str(e)})

# 删除单条日志  会返回所有的数据 ？？？


def delete_cmslog(request):
    data = json.loads(request.body.decode('utf-8'))
    print(data)
    try:
        # 根据id找到需要删除的日志记录
        obj_cmslog = Cms_log.objects.get(id=data['id'])

        obj_cmslog.delete()
        # 删除后返回
        # obj_cmslog = Cms_log.objects.all().values()
        # # 把外层的容器转为List
        # cmslog = list(obj_cmslog)
        # 返回
        return JsonResponse({'code': 1, 'msg': '批量删除日志成功'})

    except Exception as e:
        print(e)
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "删除日志时出现异常，具体错误：" + str(e)})

# 批量删除日志  会返回所有的数据 ？？？


def delete_cmslogs(request):
    # 批量删除
    data = json.loads(request.body.decode('utf-8'))
    try:
        # 根据id找到需要删除的日志记录
        for one_id in data['cmslogs']:
            obj_cmslog = Cms_log.objects.get(id=one_id['id'])
            obj_cmslog.delete()
            # 批量删除后 获取所有集合
        # obj_cmslog = Cms_log.objects.all().values()
        # # 把外层的容器转为List
        # cmslog = list(obj_cmslog)
        # 返回
        return JsonResponse({'code': 1,'msg':'单条日志删除成功'})

    except Exception as e:
        print(e)
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "批量删除日志时出现异常，具体错误：" + str(e)})


# 用户登录接口
def query_cmsusers(request):
    """查询用户信息"""
    # 接收传递过来的查询条件--- axios默认是json --- 字典类型（'inputStr'）-- data['inputStr']
    data = json.loads(request.body.decode('utf-8'))
    end_str = sha1_password(data['password'])
    try:
        # 使用ORM获取满足条件的信息 并把对象转为字典格式
        obj_cmsuers = Cms_users.objects.filter(Q(username=data['username'])
                                               & Q(password=end_str))
        if obj_cmsuers:
            username = list(obj_cmsuers.values())[0]['username']
            # 返回是否管理员
            admin = list(obj_cmsuers.values())[0]['admin']
            # 返回token
            token = generate_token(data['username'])
            return JsonResponse({'code': 1,
                                 'data': '用户验证成功',
                                 'name': username,
                                 'admin': admin,
                                 'token': token})
        else:
            return JsonResponse({'code': 0, 'data': '用户名或密码错误'})

    except Exception as e:
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "查询日志信息时出现异常，具体错误：" + str(e)})


# 用户修改密码

def modpw_cmsusers(request):
    data = json.loads(request.body.decode('utf-8'))

    end_srt = sha1_password(data['checkPass'])
    print(end_srt)
    try:
        # 根据id找到需要修改的日志记录
        obj_cmsuser = Cms_users.objects.get(username=data['username'])
        # 依次进行修改

        obj_cmsuser.password = str(end_srt)

        obj_cmsuser.save()
        # 保存后返回

        # 把外层的容器转为List

        # 返回
        return JsonResponse({'code': 1, 'data': '用户密码修改成功'})

    except Exception as e:
        print(e)
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'data': "密码修改时出现异常，具体错误：" + str(e)})


def get_random_str():
    # 获取uuid的随机数
    uuid_val = uuid.uuid4()
    # 获取uuid的随机数字符串
    uuid_str = str(uuid_val).encode('utf-8')
    # 获取md5实例
    md5 = hashlib.md5()
    # 拿取uuid的md5摘要
    md5.update(uuid_str)
    # 返回固定长度的字符串
    return md5.hexdigest()


def write_to_excel(data: list, path: str):
    """把数据库写入到Excel"""
    # 实例化一个workbook
    workbook = openpyxl.Workbook()
    # 激活一个sheet
    sheet = workbook.active
    # 为sheet命名
    sheet.title = 'cmslogs'
    # 维护表头
    tableTitle = [
        '日志ID',
        '所属公司',
        '所属部门',
        '事件类型',
        '问题描述',
        '问题归类',
        '处理办法',
        '问题来源',
        '处理人',
        '提出人',
        '登记日期',
        '解决日期',
        '状态',
        '备注']

    for col in range(len(tableTitle)):
        c = col + 1
        sheet.cell(row=1, column=c).value = tableTitle[col]
    # 准备keys
    keys = data[0].keys()
    # 准备写入数据
    for index, item in enumerate(data):
        # 遍历每一个元素
        for k, v in enumerate(keys):
            sheet.cell(row=index + 2, column=k + 1, value=str(item[v]))
    # 写入到文件
    workbook.save(path)


def export_comslogs_excel(request):
    """到处数据到excel"""
    # # 获取所有的学生信息
    # obj_cms_log = Cms_log.objects.all().values()
    # # 转为List
    # cmslogs = list(obj_cms_log)
    # 准备名称
    excel_name = get_random_str() + ".xlsx"
    # 准备写入的路劲
    path = os.path.join(settings.MEDIA_ROOT, excel_name)
    # 写入到Excel
    global export_list
    print(export_list)
    write_to_excel(export_list, path)

    # 返回
    return JsonResponse({'code': 1, 'name': excel_name})

# 近7天每人新增日志量
def info_a(request):
    SQL_str = '''SELECT days.handler, nvl(m.aa, 0)
  FROM (SELECT distinct handler as handler
          FROM cms_log_hdr
        connect BY LEVEL <= 7) days
  LEFT JOIN (select handler as isday, count(id) as aa
               from cms_log_hdr
              where createdate > = sysdate - 6
                and createdate <= sysdate
              group by handler
              order by count(id)) m
    ON days.handler = m.isday'''
    cursor = connection.cursor()
    cursor.execute(SQL_str)
    # 返回时元组转换成字典
    info_a = dict(cursor.fetchall())

    keys = [key for key, value in info_a.items()]
    values = [value for key, value in info_a.items()]
    # name =np.array(keys)
    # num=np.array(values)

    return JsonResponse({'name': keys, 'num': values})



# 近7天来源系统 日志情况
def info_b(request):
  #   SQL_str = '''SELECT days.event_system, nvl(m.aa, 0)
  # FROM (SELECT distinct event_system as event_system
  #         FROM cms_log_hdr
  #       connect BY LEVEL <= 7) days
  # LEFT JOIN (select event_system as isday, count(id) as aa
  #              from cms_log_hdr
  #             where createdate > = sysdate - 6
  #               and createdate <= sysdate
  #             group by event_system
  #             order by count(id)) m
  #   ON days.event_system = m.isday'''


    SQL_str = '''select event_system as isday, count(id) as aa
  from cms_log_hdr
 where createdate > = sysdate - 7
   and createdate <= sysdate
 group by event_system
 order by count(id)'''

    cursor = connection.cursor()
    cursor.execute(SQL_str)
    # 返回时元组转换成字典
    info_b = cursor.fetchall()


    list = []
    list1 = []

    for k, v in info_b:
        end_dict = {}

        end_dict['value'] = v
        print(k)
        end_dict['name'] = k

        list.append(end_dict)
        list1.append(str(k))
    print(list1)
    print(list)
    # keys = [key for key, value in info_a.items()]
    # values = [value for key, value in info_a.items()]
    # name =np.array(keys)
    # num=np.array(values)



    return JsonResponse({'name':list1,'list' :list})




# 所有公司 日志情况
def info_d(request):
    SQL_str = ''' select ownername, count(id) as aa
      from cms_log_hdr
     group by ownername
     order by count(id)'''
    cursor = connection.cursor()
    cursor.execute(SQL_str)
    # 返回时元组转换成字典
    info_b = cursor.fetchall()


    list = []
    list1 = []

    for k, v in info_b:
        end_dict = {}

        end_dict['value'] = v
        print(k)
        end_dict['name'] = k

        list.append(end_dict)
        list1.append(str(k))
    print(list1)
    print(list)
    # keys = [key for key, value in info_a.items()]
    # values = [value for key, value in info_a.items()]
    # name =np.array(keys)
    # num=np.array(values)



    return JsonResponse({'name':list1,'list' :list})


# 近7天日增量
def info_c(request):
    SQL_str = '''SELECT days.today, nvl(m.aa, 0)
  FROM (SELECT to_char(SYSDATE - LEVEL + 1, 'yyyy-mm-dd') today
          FROM DUAL
        connect BY LEVEL <= 7) days
  LEFT JOIN (select to_char(createdate, 'yyyy-mm-dd') as isday,
                    count(id) as aa
               from cms_log_hdr
              where createdate > = sysdate - 6
                and createdate <= sysdate
              group by to_char(createdate, 'yyyy-mm-dd')
              order by to_char(createdate, 'yyyy-mm-dd')) m
    ON days.today = m.isday order by  days.today'''
    cursor = connection.cursor()
    cursor.execute(SQL_str)
    # 返回时元组转换成字典
    info_c = dict(cursor.fetchall())
    print(info_c)
    keys = [key for key, value in info_c.items()]
    values = [value for key, value in info_c.items()]
    # name =np.array(keys)
    # num=np.array(values)
    print(keys)
    print(values)
    return JsonResponse({'date': keys, 'num': values})


# 需求 人为原因 系统原因 占比
def info_e_end_data(str, type_list, i):
    # 事项数目 data
    data_list = []

    # 最终字典
    one_dict = {}
    for k in str:
        one_dict['name'] = type_list[i]
        data_list.append(k[i + 1])
    one_dict['data'] = data_list

    return one_dict


def info_e(request):
    SQL_str = '''select qa.event_system,
       sum(nvl(decode(qa.event_type, '需求', qa.num),0)) 需求,
       sum(nvl(decode(qa.event_type, '人为原因', qa.num),0)) 人为原因,
              sum(nvl(decode(qa.event_type, '系统原因', qa.num),0))系统原因,
                   sum(nvl(decode(qa.event_type, '个人计划', qa.num),0)) 个人计划,  
                      sum(nvl(decode(qa.event_type, '其他', qa.num),0)) 其他 

  from (select a.event_system, a.event_type, count(*) as num
          from cms_log_hdr a
          where a.createdate >= trunc(sysdate,'iw') - 7
          and a.createdate <=  trunc(sysdate,'iw') - 1
         group by a.event_system, a.event_type) qa
         group by  qa.event_system order by  qa.event_system'''
    cursor = connection.cursor()
    cursor.execute(SQL_str)
    # 返回时元组转换成字典
    info_e = cursor.fetchall()

    # 生成系统名称list
    system_list = []
    for p in info_e:
        system_list.append(p[0])
    # print(system_list)

    # 系统名称 下级 事件分类菜单
    type_list = ['需求', '人为原因', '系统原因', '个人计划', '其他']

    data_list = []
    for i in range(len(type_list)):
        data_list.append(info_e_end_data(info_e, type_list, i))

    info_e_end_dict = {}
    info_e_end_dict['system_list'] = system_list
    info_e_end_dict['type_list'] = type_list
    info_e_end_dict['data'] = data_list
    return JsonResponse(info_e_end_dict)